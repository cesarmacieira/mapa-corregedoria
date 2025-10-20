import io
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import streamlit as st
from io import BytesIO
from datetime import datetime
import plotly.express as px
import matplotlib.pyplot as plt
import requests

st.set_page_config(page_title="JFCE",
                   page_icon="chart",
                   layout="wide",
                   initial_sidebar_state="auto",
                   menu_items=None)

# Extração dos dados
url = "https://webservice-d.jfce.jus.br/sarh_new/json/buscarMapaCorregedoria"
try:
    resposta = requests.get(url, timeout=20)
    resposta.raise_for_status()
    dados_json = resposta.json()
    funcionarios = dados_json["mapaCorregedoria"]["funcionarios"]
    dados = pd.DataFrame(funcionarios)

except requests.exceptions.RequestException as e:
    st.markdown(
        f"""
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
                    height:90vh;text-align:center;color:#b00020;">
            <img src="https://cdn-icons-png.flaticon.com/512/564/564619.png" width="120" style="margin-bottom:20px;">
            <h2 style="font-family:'Segoe UI',sans-serif;">Erro ao acessar o webservice</h2>
            <p style="max-width:600px;font-size:18px;">
                404 - Serviço não encontrado ou temporariamente indisponível.<br><br>
                <b>Detalhe técnico:</b> {e}<br><br>
                <span style="font-size:16px;color:#444;">Comunique a equipe responsável.</span>
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.stop()

except KeyError:
    st.markdown(
        """
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
                    height:90vh;text-align:center;color:#b00020;">
            <img src="https://cdn-icons-png.flaticon.com/512/564/564619.png" width="120" style="margin-bottom:20px;">
            <h2 style="font-family:'Segoe UI',sans-serif;">Erro no formato dos dados</h2>
            <p style="max-width:600px;font-size:18px;">
                O webservice retornou um formato inesperado.<br><br>
                <span style="font-size:16px;color:#444;">Comunique a equipe técnica responsável.</span>
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.stop()

except Exception as e:
    st.markdown(
        f"""
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
                    height:90vh;text-align:center;color:#b00020;">
            <img src="https://cdn-icons-png.flaticon.com/512/564/564619.png" width="120" style="margin-bottom:20px;">
            <h2 style="font-family:'Segoe UI',sans-serif;">Erro inesperado</h2>
            <p style="max-width:600px;font-size:18px;">
                Algo deu errado durante a execução.<br><br>
                <b>Detalhe técnico:</b> {e}<br><br>
                <span style="font-size:16px;color:#444;">Por favor, comunique a equipe responsável.</span>
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.stop()

# Transformação dos dados
## Definição do vínculo
def definir_vinculo(row):
    if row["divisao"] == "SERVIDOR DO QUADRO":
        return "EFETIVO"
    elif row["divisao"] == "SEM VINCULO":
        return "SEM VÍNCULO"
    elif row["divisao"] == "SERVIDOR DE OUTROS ORGAOS":
        if row["provimento"] == "AUTORIZACAO PARA EXERCICIO PROVISORIO":
            return "EXERCÍCIO PROVISÓRIO"
        elif "REMOV" in row["situacao"]:
            return "REMOVIDO"
        else:
            return "REQUISITADO"
    else:
        return "OUTRO"
dados["vinculo"] = dados.apply(definir_vinculo, axis=1)

## Ordenação dos cargos e tratamento de valores ausentes
cargo_order = [
    "ANALISTA JUDICIÁRIO/ ADMINISTRATIVA",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (ANÁLISE DE SISTEMAS DE INFORMAÇÃO)",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (BIBLIOTECONOMIA)",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (CONTABILIDADE)",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (ENGENHARIA (CIVIL))",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (INFORMÁTICA (INFRAESTRUTURA))",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (MEDICINA (CLÍNICA GERAL))",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (PSICOLOGIA)",
    "ANALISTA JUDICIÁRIO/ APOIO ESPECIALIZADO (TECNOLOGIA DA INFORMAÇÃO)",
    "ANALISTA JUDICIÁRIO/ JUDICIÁRIA",
    "ANALISTA JUDICIÁRIO/ JUDICIÁRIA (OFICIAL DE JUSTIÇA AVALIADOR FEDERAL)",
    "TÉCNICO JUDICIÁRIO/ ADMINISTRATIVA",
    "TÉCNICO JUDICIÁRIO/ ADMINISTRATIVA (AGENTE DE POLÍCIA JUDICIAL)",
    "TÉCNICO JUDICIÁRIO/ APOIO ESPECIALIZADO (CONTABILIDADE)",
    "TÉCNICO JUDICIÁRIO/ APOIO ESPECIALIZADO (TECNOLOGIA DA INFORMAÇÃO)",
    "SEM DESCRIÇÃO DE CARGO"]
dados['descricaoCargo'] = dados['descricaoCargo'].apply(lambda x: "SEM DESCRIÇÃO DE CARGO" if pd.isna(x) or x == "" else x)
dados['descricaoCargo'] = pd.Categorical(dados['descricaoCargo'], categories=cargo_order, ordered=True)

## Substituição de nomes de lotações (gabinetes por diretoria/subdiretorias)
def substituir_gabinete(lotacao):
    if lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DO FORO":
        return "DIRETORIA DO FORO"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE LIMOEIRO DO NORTE":
        return "SUBDIRETORIA DO FORO - LIMOEIRO DO NORTE"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE JUAZEIRO DO NORTE":
        return "SUBDIRETORIA DO FORO - JUAZEIRO DO NORTE"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE SOBRAL":
        return "SUBDIRETORIA DO FORO - SOBRAL"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE CRATEUS":
        return "SUBDIRETORIA DO FORO - CRATEÚS"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE QUIXADA":
        return "SUBDIRETORIA DO FORO - QUIXADÁ"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE TAUA":
        return "SUBDIRETORIA DO FORO - TAUÁ"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE IGUATU":
        return "SUBDIRETORIA DO FORO - IGUATU"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE ITAPIPOCA":
        return "SUBDIRETORIA DO FORO - ITAPIPOCA"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO-MARACANAU-CE":
        return "SUBDIRETORIA DO FORO - MARACANAÚ"
    else:
        return lotacao
dados['descricaoLotacaoPai_subdiretoria'] = dados['descricaoLotacaoPai'].apply(substituir_gabinete)

## Simplificação das lotações para agrupamento (diretoria, subdiretorias, turmas recursais, núcleos)
def simplificar_lotacao(lotacao):
    if lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DO FORO":
        return "DIRETORIA DO FORO"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE LIMOEIRO DO NORTE":
        return "SUBDIRETORIA DO FORO - LIMOEIRO DO NORTE"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE JUAZEIRO DO NORTE":
        return "SUBDIRETORIA DO FORO - JUAZEIRO DO NORTE"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE SOBRAL":
        return "SUBDIRETORIA DO FORO - SOBRAL"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE CRATEUS":
        return "SUBDIRETORIA DO FORO - CRATEUS"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE QUIXADA":
        return "SUBDIRETORIA DO FORO - QUIXADÁ"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE TAUA":
        return "SUBDIRETORIA DO FORO - TAUÁ"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE IGUATU":
        return "SUBDIRETORIA DO FORO - IGUATU"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO DE ITAPIPOCA":
        return "SUBDIRETORIA DO FORO - ITAPIPOCA"
    elif lotacao == "GABINETE DO JUIZ FEDERAL DIRETOR DA SUBSEÇAO-MARACANAU-CE":
        return "SUBDIRETORIA DO FORO - MARACANAÚ"
    elif lotacao in ["1ª TURMA RECURSAL", "2ª TURMA RECURSAL", "3ª TURMA RECURSAL"]:
        return "TURMAS RECURSAIS"
    elif lotacao in [
        "NUCLEO DE AUDITORIA INTERNA", "NUCLEO DE ESTRATEGIA, GOVERNANÇA E INTEGRIDADE",
        "NUCLEO DE GESTAO DE PESSOAS", "NUCLEO DE GESTAO ORCAMENTARIA FINANCEIRA CONTABIL E PATRIMONIAL",
        "NUCLEO DE INFRAESTRUTURA E ADMINISTRAÇAO PREDIAL", "NUCLEO DE INTELIGENCIA, SEGURANÇA E TRANSPORTE",
        "NUCLEO DE TECNOLOGIA DA INFORMAÇAO E COMUNICAÇAO", "NUCLEO JUDICIARIO"]:
        return "NUCLEO"
    else:
        return lotacao
dados['lotacaoSimplificada'] = dados['descricaoLotacaoPai'].apply(simplificar_lotacao)

## Definição dos grupos de subsecao para somatórios (totalizadores)
grupos_subsecao = {
    'TOTAL FORTALEZA': [
        "1ª VARA - FORTALEZA-CE","2ª VARA - FORTALEZA-CE","3ª VARA - FORTALEZA-CE","4ª VARA - FORTALEZA-CE",
        "5ª VARA - FORTALEZA-CE","6ª VARA - FORTALEZA-CE","7ª VARA - FORTALEZA-CE","8ª VARA - FORTALEZA-CE",
        "9ª VARA - FORTALEZA-CE","10ª VARA - FORTALEZA-CE","11ª VARA - FORTALEZA-CE","12ª VARA - FORTALEZA-CE",
        "13ª VARA - JEF - FORTALEZA-CE","14ª VARA - JEF - FORTALEZA-CE","20ª VARA - FORTALEZA-CE",
        "21ª VARA - JEF - FORTALEZA-CE","26ª VARA - JEF - FORTALEZA - CE","28ª VARA - JEF - FORTALEZA-CE",
        "32ª VARA - FORTALEZA-CE","33ª VARA - FORTALEZA-CE",
        "1ª TURMA RECURSAL","2ª TURMA RECURSAL","3ª TURMA RECURSAL",
        "SECRETARIA ADMINISTRATIVA","DIRETORIA DO FORO"],
    "TOTAL RECURSAIS": ["1ª TURMA RECURSAL","2ª TURMA RECURSAL","3ª TURMA RECURSAL"],
    'TOTAL LIMOEIRO DO NORTE': ["15ª VARA - LIMOEIRO DO NORTE-CE","29ª VARA - JEF - LIMOEIRO DO NORTE - CE",
                        "SUBDIRETORIA DO FORO - LIMOEIRO DO NORTE"],
    'TOTAL JUAZEIRO DO NORTE': ["16ª VARA - JUAZEIRO DO NORTE-CE","17ª VARA - JEF - JUAZEIRO DO NORTE-CE",
                        "30ª VARA - JEF - JUAZEIRO DO NORTE - CE",
                        "SUBDIRETORIA DO FORO - JUAZEIRO DO NORTE"],
    'TOTAL SOBRAL': ["18ª VARA - SOBRAL-CE","19ª VARA - JEF - SOBRAL-CE","31ª VARA - JEF - SOBRAL - CE",
                    "SUBDIRETORIA DO FORO - SOBRAL"],
    'TOTAL CRATEÚS': ["22ª VARA - CRATEÚS-CE","SUBDIRETORIA DO FORO - CRATEÚS"],
    'TOTAL QUIXADÁ': ["23ª VARA - QUIXADÁ-CE","SUBDIRETORIA DO FORO - QUIXADÁ"],
    'TOTAL TAUÁ': ["24ª VARA - TAUÁ-CE","SUBDIRETORIA DO FORO - TAUÁ"],
    'TOTAL IGUATU': ["25ª VARA - IGUATU-CE","SUBDIRETORIA DO FORO - IGUATU"],
    'TOTAL ITAPIPOCA': ["27ª VARA- ITAPIPOCA-CE","SUBDIRETORIA DO FORO - ITAPIPOCA"],
    'TOTAL MARACANAÚ': ["34ª VARA - MARACANAÚ-CE","35ª VARA - JEF - MARACANAÚ-CE",
                        "SUBDIRETORIA DO FORO - MARACANAÚ"],
    'TOTAL SERVIDORES COM LOTAÇÃO': ['TOTAL FORTALEZA','TOTAL LIMOEIRO DO NORTE''TOTAL JUAZEIRO DO NORTE','TOTAL SOBRAL','TOTAL CRATEÚS',
                        'TOTAL QUIXADÁ','TOTAL TAUÁ','TOTAL IGUATU','TOTAL ITAPIPOCA', 'TOTAL MARACANAÚ'],
    'TOTAL SERVIDORES SEM LOTAÇÃO': ["SERVIDORA EM LICENÇA GESTANTE EXERC. FUNÇÃO",
                    "SERVIDORES CEDIDOS/EXERCICIO PROVISÓRIO/REMOVIDO"],
    'TOTAL SERVIDORES': ['TOTAL SERVIDORES COM LOTAÇÃO','TOTAL SERVIDORES SEM LOTAÇÃO'],
    'TOTAL NÚCLEOS': ["NUCLEO DE AUDITORIA INTERNA","NUCLEO DE ESTRATEGIA, GOVERNANÇA E INTEGRIDADE",
                    "NUCLEO DE GESTAO DE PESSOAS","NUCLEO DE GESTAO ORCAMENTARIA FINANCEIRA CONTABIL E PATRIMONIAL",
                    "NUCLEO DE INFRAESTRUTURA E ADMINISTRAÇAO PREDIAL","NUCLEO DE INTELIGENCIA, SEGURANÇA E TRANSPORTE",
                    "NUCLEO DE TECNOLOGIA DA INFORMAÇAO E COMUNICAÇAO","NUCLEO JUDICIARIO"],
}

## Definição da ordem das colunas e abas
ordem_colunas_mapa  = [
    "1ª VARA - FORTALEZA-CE","2ª VARA - FORTALEZA-CE","3ª VARA - FORTALEZA-CE","4ª VARA - FORTALEZA-CE",
    "5ª VARA - FORTALEZA-CE","6ª VARA - FORTALEZA-CE","7ª VARA - FORTALEZA-CE","8ª VARA - FORTALEZA-CE",
    "9ª VARA - FORTALEZA-CE","10ª VARA - FORTALEZA-CE","11ª VARA - FORTALEZA-CE","12ª VARA - FORTALEZA-CE",
    "13ª VARA - JEF - FORTALEZA-CE","14ª VARA - JEF - FORTALEZA-CE","20ª VARA - FORTALEZA-CE",
    "21ª VARA - JEF - FORTALEZA-CE","26ª VARA - JEF - FORTALEZA - CE","28ª VARA - JEF - FORTALEZA-CE",
    "32ª VARA - FORTALEZA-CE","33ª VARA - FORTALEZA-CE",
    "1ª TURMA RECURSAL","2ª TURMA RECURSAL","3ª TURMA RECURSAL","TOTAL RECURSAIS",
    "SECRETARIA ADMINISTRATIVA","DIRETORIA DO FORO",
    "TOTAL FORTALEZA",

    "15ª VARA - LIMOEIRO DO NORTE-CE","29ª VARA - JEF - LIMOEIRO DO NORTE - CE",
    "SUBDIRETORIA DO FORO - LIMOEIRO DO NORTE",'TOTAL LIMOEIRO DO NORTE',
    
    "16ª VARA - JUAZEIRO DO NORTE-CE","17ª VARA - JEF - JUAZEIRO DO NORTE-CE",
    "30ª VARA - JEF - JUAZEIRO DO NORTE - CE",
    "SUBDIRETORIA DO FORO - JUAZEIRO DO NORTE",'TOTAL JUAZEIRO DO NORTE',
    
    "18ª VARA - SOBRAL-CE","19ª VARA - JEF - SOBRAL-CE","31ª VARA - JEF - SOBRAL - CE",
    "SUBDIRETORIA DO FORO - SOBRAL",'TOTAL SOBRAL',
    
    "22ª VARA - CRATEÚS-CE","SUBDIRETORIA DO FORO - CRATEÚS",'TOTAL CRATEÚS',

    "23ª VARA - QUIXADÁ-CE","SUBDIRETORIA DO FORO - QUIXADÁ",'TOTAL QUIXADÁ',

    "24ª VARA - TAUÁ-CE","SUBDIRETORIA DO FORO - TAUÁ",'TOTAL TAUÁ',

    "25ª VARA - IGUATU-CE","SUBDIRETORIA DO FORO - IGUATU",'TOTAL IGUATU',

    "27ª VARA- ITAPIPOCA-CE","SUBDIRETORIA DO FORO - ITAPIPOCA",'TOTAL ITAPIPOCA',

    "34ª VARA - MARACANAÚ-CE","35ª VARA - JEF - MARACANAÚ-CE",
    "SUBDIRETORIA DO FORO - MARACANAÚ",'TOTAL MARACANAÚ',
    
    'TOTAL SERVIDORES COM LOTAÇÃO',

    "SERVIDORA EM LICENÇA GESTANTE EXERC. FUNÇÃO","SERVIDORES CEDIDOS/EXERCICIO PROVISÓRIO/REMOVIDO",
    'TOTAL SERVIDORES SEM LOTAÇÃO',
    'TOTAL SERVIDORES',
    
    "NUCLEO DE AUDITORIA INTERNA","NUCLEO DE ESTRATEGIA, GOVERNANÇA E INTEGRIDADE",
    "NUCLEO DE GESTAO DE PESSOAS","NUCLEO DE GESTAO ORCAMENTARIA FINANCEIRA CONTABIL E PATRIMONIAL",
    "NUCLEO DE INFRAESTRUTURA E ADMINISTRAÇAO PREDIAL","NUCLEO DE INTELIGENCIA, SEGURANÇA E TRANSPORTE",
    "NUCLEO DE TECNOLOGIA DA INFORMAÇAO E COMUNICAÇAO","NUCLEO JUDICIARIO","TOTAL NÚCLEOS"]

ordem_abas_arquivo = [
    "SECRETARIA ADMINISTRATIVA","DIRETORIA DO FORO",
    "1ª TURMA RECURSAL","2ª TURMA RECURSAL","3ª TURMA RECURSAL",
    "1ª VARA - FORTALEZA-CE","2ª VARA - FORTALEZA-CE","3ª VARA - FORTALEZA-CE","4ª VARA - FORTALEZA-CE",
    "5ª VARA - FORTALEZA-CE","6ª VARA - FORTALEZA-CE","7ª VARA - FORTALEZA-CE","8ª VARA - FORTALEZA-CE",
    "9ª VARA - FORTALEZA-CE","10ª VARA - FORTALEZA-CE","11ª VARA - FORTALEZA-CE","12ª VARA - FORTALEZA-CE",
    "13ª VARA - JEF - FORTALEZA-CE","14ª VARA - JEF - FORTALEZA-CE","20ª VARA - FORTALEZA-CE",
    "21ª VARA - JEF - FORTALEZA-CE","26ª VARA - JEF - FORTALEZA - CE","28ª VARA - JEF - FORTALEZA-CE",
    "32ª VARA - FORTALEZA-CE","33ª VARA - FORTALEZA-CE",

    "15ª VARA - LIMOEIRO DO NORTE-CE","29ª VARA - JEF - LIMOEIRO DO NORTE - CE",
    "SUBDIRETORIA DO FORO - LIMOEIRO DO NORTE",
    
    "16ª VARA - JUAZEIRO DO NORTE-CE","17ª VARA - JEF - JUAZEIRO DO NORTE-CE",
    "30ª VARA - JEF - JUAZEIRO DO NORTE - CE",
    "SUBDIRETORIA DO FORO - JUAZEIRO DO NORTE",
    
    "18ª VARA - SOBRAL-CE","19ª VARA - JEF - SOBRAL-CE","31ª VARA - JEF - SOBRAL - CE",
    "SUBDIRETORIA DO FORO - SOBRAL",
    
    "22ª VARA - CRATEÚS-CE","SUBDIRETORIA DO FORO - CRATEÚS",

    "23ª VARA - QUIXADÁ-CE","SUBDIRETORIA DO FORO - QUIXADÁ",

    "24ª VARA - TAUÁ-CE","SUBDIRETORIA DO FORO - TAUÁ",'TOTAL TAUÁ',

    "25ª VARA - IGUATU-CE","SUBDIRETORIA DO FORO - IGUATU",

    "27ª VARA- ITAPIPOCA-CE","SUBDIRETORIA DO FORO - ITAPIPOCA",

    "34ª VARA - MARACANAÚ-CE","35ª VARA - JEF - MARACANAÚ-CE",
    "SUBDIRETORIA DO FORO - MARACANAÚ",

    "SERVIDORA EM LICENÇA GESTANTE EXERC. FUNÇÃO","SERVIDORES CEDIDOS/EXERCICIO PROVISÓRIO/REMOVIDO",
    
    "NUCLEO DE AUDITORIA INTERNA","NUCLEO DE ESTRATEGIA, GOVERNANÇA E INTEGRIDADE",
    "NUCLEO DE GESTAO DE PESSOAS","NUCLEO DE GESTAO ORCAMENTARIA FINANCEIRA CONTABIL E PATRIMONIAL",
    "NUCLEO DE INFRAESTRUTURA E ADMINISTRAÇAO PREDIAL","NUCLEO DE INTELIGENCIA, SEGURANÇA E TRANSPORTE",
    "NUCLEO DE TECNOLOGIA DA INFORMAÇAO E COMUNICAÇAO","NUCLEO JUDICIARIO"]

## Função para limpar nomes de abas com caracteres inválidos
def limpar_nome_aba(nome):
    return re.sub(r'[\[\]\:\*\?\/\\]', '', nome)[:31]

# Construção do Streamlit
## Sidebar e seleção de painéis
painel = st.sidebar.radio("Selecione o painel:", ["Mapa da Corregedoria", "Análises", "Dados Brutos"])

## Painel Mapa da Corregedoria
if painel == "Mapa da Corregedoria":

    ### Título e descrição
    st.title("Mapa da Corregedoria")
    st.write('Versão 1.0.0')
    st.write("Este é o mapa da corregedoria da Justiça Federal do Ceará.")

    ### Construção das tabelas de contingência
    #### Tabelas de contingência iniciais
    ct_cargo = pd.crosstab(dados['descricaoCargo'], dados['descricaoLotacaoPai_subdiretoria'])
    ct_vinculo = pd.crosstab(dados['vinculo'], dados['descricaoLotacaoPai_subdiretoria'])

    #### Definição dos nomes dos índices
    ct_cargo.index.name = "CARGO"
    ct_vinculo.index.name = "VÍNCULO"

    #### Cálculo dos totais conforme grupos de subseção
    ct_cargo.columns = ct_cargo.columns.str.strip().str.upper()
    ct_vinculo.columns = ct_vinculo.columns.str.strip().str.upper()

    #### Adição das colunas de totalizadores
    for total_col, subcols in grupos_subsecao.items():
        subcols_upper = [s.strip().upper() for s in subcols]
        colunas_existentes_cargo = [c for c in ct_cargo.columns if c in subcols_upper]
        colunas_existentes_vinculo = [c for c in ct_vinculo.columns if c in subcols_upper]
        ct_cargo[total_col] = ct_cargo[colunas_existentes_cargo].sum(axis=1) if colunas_existentes_cargo else 0
        ct_vinculo[total_col] = ct_vinculo[colunas_existentes_vinculo].sum(axis=1) if colunas_existentes_vinculo else 0

    #### Reordenação das colunas conforme ordem definida
    colunas_finais_cargo = [c for c in ordem_colunas_mapa if c in ct_cargo.columns]
    colunas_finais_vinculo = [c for c in ordem_colunas_mapa if c in ct_vinculo.columns]

    #### Reordenação das tabelas de contingência
    ct_cargo = ct_cargo[colunas_finais_cargo + [c for c in ct_cargo.columns if c not in colunas_finais_cargo]].fillna(0)
    ct_vinculo = ct_vinculo[colunas_finais_vinculo + [c for c in ct_vinculo.columns if c not in colunas_finais_vinculo]].fillna(0)

    #### Adição da linha de totais gerais
    ct_cargo.loc["TOTAL"] = ct_cargo.sum(numeric_only=True)
    ct_vinculo.loc["TOTAL"] = ct_vinculo.sum(numeric_only=True)

    #### Preenchimento de valores ausentes com zero
    ct_cargo = ct_cargo.fillna(0)
    ct_vinculo = ct_vinculo.fillna(0)

    #### Exibição das tabelas no Streamlit
    st.write("Cargos por serventia")
    st.dataframe(ct_cargo)
    st.write("Provimentos por serventia")
    st.dataframe(ct_vinculo)

    #### Criação do MultiIndex
    ct_cargo['VARIÁVEL'] = 'CARGO'
    ct_vinculo['VARIÁVEL'] = 'VÍNCULO'

    #### Concatenação das tabelas
    mapa_corregedoria = pd.concat([ct_cargo, ct_vinculo])
    mapa_corregedoria = mapa_corregedoria.reset_index().rename(columns={'index': 'CATEGORIA'})
    mapa_corregedoria = mapa_corregedoria.set_index(['VARIÁVEL', mapa_corregedoria.columns[0]])

    #### Exportação para Excel com bordas
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        mapa_corregedoria.to_excel(writer, sheet_name='Mapa da Corregedoria', merge_cells=True)
    buffer.seek(0)

    # Abre o arquivo gerado
    wb = load_workbook(buffer)
    ws = wb['Mapa da Corregedoria']

    # Define o estilo de borda fina preta
    thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))

    # Aplica bordas em todas as células preenchidas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    #### Nome do arquivo com data
    hoje = datetime.today().strftime('%d-%m-%Y')
    nome_arquivo = f'Mapa da Corregedoria {hoje}.xlsx'

    #### Botão de download
    buffer_multi = io.BytesIO()
    with pd.ExcelWriter(buffer_multi, engine='openpyxl') as writer:
        mapa_corregedoria.to_excel(writer, sheet_name='Mapa da Corregedoria', merge_cells=True)
        for aba in ordem_abas_arquivo:
            df_aba = dados[dados['descricaoLotacaoPai_subdiretoria'].astype(str).str.strip() == aba]
            df_aba = df_aba.drop(columns=['descricaoLotacaoPai_subdiretoria','lotacaoSimplificada'], errors='ignore')
            aba_limpa = re.sub(r'[\\/*?:\[\]]', ' ', aba).strip()[:31]
            df_aba.to_excel(writer, index=False, sheet_name=aba_limpa)
    buffer_multi.seek(0)
    nome_arquivo_abas = f'Mapa da Corregedoria {hoje}.xlsx'
    st.download_button(label="📥 Download Mapa da Corregedoria", data=buffer_multi, 
        file_name=nome_arquivo_abas, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    ### Seção de lotações e provimento       
    st.title("Lotações")
    lotacoes_validas = [s for s in ordem_abas_arquivo if s in dados['descricaoLotacaoPai_subdiretoria'].unique()]
    selecao_lotacoes = st.selectbox("Selecione a lotação para ver a quantidade de servidores em uma lotação específica:", lotacoes_validas)

    tabela_lotacoes = dados[dados['descricaoLotacaoPai_subdiretoria'] == selecao_lotacoes]
    tabela_lotacoes = (tabela_lotacoes.groupby("descricaoCargo").size().reset_index(name='QUANTIDADE DE SERVIDORES')
        .sort_values(by='QUANTIDADE DE SERVIDORES', ascending=False))
    tabela_lotacoes = tabela_lotacoes.reset_index(drop=True)
    st.dataframe(tabela_lotacoes.rename(columns={"descricaoCargo": "CARGO"}), use_container_width=False)
    
    hoje = datetime.today().strftime('%d-%m-%Y')
    nome_arquivo = f"Lotação - {selecao_lotacoes} - {hoje}.xlsx"
    tabela_lotacoes.to_excel(nome_arquivo, index=False, engine='openpyxl')

    # Cria botão de download
    # with open(nome_arquivo, "rb") as file:
    #     st.download_button(label=f"📥 Download dados da {selecao_lotacoes}", data=file,
    #         file_name=nome_arquivo, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #         key="download_lotacoes")
        
    st.title("Provimento")
    selecao_provimento = st.selectbox("Selecione a lotação para ver a quantidade de servidores em uma lotação específica:", 
                                      lotacoes_validas,key="select_lotacao_provimento")

    tabela_provimento = dados[dados['descricaoLotacaoPai'] == selecao_provimento]
    tabela_provimento = (tabela_provimento.groupby("vinculo").size().reset_index(name='QUANTIDADE DE SERVIDORES')
        .sort_values(by='QUANTIDADE DE SERVIDORES', ascending=False))
    tabela_provimento = tabela_provimento.reset_index(drop=True)
    st.dataframe(tabela_provimento.rename(columns={"vinculo": "VÍNCULO"}), use_container_width=False)

    hoje = datetime.today().strftime('%d-%m-%Y')
    nome_arquivo = f"Lotação - {selecao_provimento} - {hoje}.xlsx"
    tabela_provimento.to_excel(nome_arquivo, index=False, engine='openpyxl')

    # Cria botão de download
    # with open(nome_arquivo, "rb") as file:
    #     st.download_button(label=f"📥 Download dados da {selecao_provimento}", data=file,
    #         file_name=nome_arquivo, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #         key="download_provimento")

## Painel Análises       
elif painel == "Análises":
    st.title("Análises")
    st.write("Explore os dados de servidores por lotação, cargo e vínculo.")

    #### Filtros
    colf1, colf2, colf3 = st.columns(3)
    with colf1:
        filtro_lotacao = st.multiselect("Selecione a lotação:", options=sorted(dados['descricaoLotacaoPai_subdiretoria'].dropna().unique()), 
                                        default=None, placeholder="Nenhuma lotação selecionada")
    with colf2:
        filtro_cargo = st.multiselect("Selecione o cargo:", options=sorted(dados['descricaoCargo'].dropna().unique()), 
                                       default=None, placeholder="Nenhum cargo selecionado")
    with colf3:
        filtro_vinculo = st.multiselect("Selecione o vínculo:", options=sorted(dados['vinculo'].dropna().unique()), 
                                         default=None, placeholder="Nenhum vínculo selecionado")

    #### Aplicação dos filtros
    df_filtrado = dados.copy()
    if filtro_lotacao:
        df_filtrado = df_filtrado[df_filtrado['descricaoLotacaoPai_subdiretoria'].isin(filtro_lotacao)]
    if filtro_cargo:
        df_filtrado = df_filtrado[df_filtrado['descricaoCargo'].isin(filtro_cargo)]
    if filtro_vinculo:
        df_filtrado = df_filtrado[df_filtrado['vinculo'].isin(filtro_vinculo)]

    #### Agregações
    lotacoes_df = df_filtrado['descricaoLotacaoPai_subdiretoria'].value_counts().sort_values(ascending=False).reset_index()
    lotacoes_df.columns = ['Lotação', 'Servidores']
    cargos_df = df_filtrado['descricaoCargo'].value_counts().sort_values(ascending=False).reset_index()
    cargos_df.columns = ['Cargo', 'Servidores']
    vinculo_df = df_filtrado['vinculo'].value_counts().sort_values(ascending=False).reset_index()
    vinculo_df.columns = ['Vínculo', 'Servidores']

    # opcao_top = st.selectbox("Selecione o número de categorias para exibir:", [10, 20, 30, 50, len(lotacoes_df)], index=1)
    # dados_filtrados = lotacoes_df.nlargest(opcao_top, 'Servidores')
    # altura = 800
    # fig1 = px.bar(dados_filtrados, y='Lotação', x='Servidores', orientation='h', color='Servidores', color_continuous_scale='tealgrn', title=f"Lotações com mais servidores", text='Servidores')
    # fig1.update_layout(yaxis={'categoryorder':'total ascending','tickfont':dict(size=16)}, xaxis={'tickfont':dict(size=16)}, title={'font':dict(size=24)}, height=altura, bargap=0.4)
    # fig1.update_traces(texttemplate='%{x}', textposition='outside', cliponaxis=False)
    # st.plotly_chart(fig1, use_container_width=True)

    st.markdown("### Lotações com mais servidores")
    opcao_top = st.selectbox("Selecione o número de categorias para exibir:", [5, 10, 20, 30, 50, len(lotacoes_df)], index=1)
    dados_filtrados = lotacoes_df.nlargest(opcao_top, 'Servidores')
    fig1 = px.bar(dados_filtrados, y='Lotação', x='Servidores', orientation='h', color='Servidores', color_continuous_scale='tealgrn', title=None, text='Servidores')
    fig1.update_layout(yaxis={'categoryorder':'total ascending','tickfont':dict(size=16)}, xaxis={'tickfont':dict(size=16)}, bargap=0.4, margin=dict(t=20))
    fig1.update_traces(texttemplate='%{x}', textposition='outside', cliponaxis=False)
    st.plotly_chart(fig1, use_container_width=True)

    col1, col2 = st.columns([2,1])
    with col1:
        st.markdown("### Cargos por servidores")
        fig2 = px.bar(cargos_df, y='Cargo', x='Servidores', orientation='h', color='Servidores', color_continuous_scale='tealgrn', title=None, text='Servidores')
        fig2.update_layout(yaxis={'categoryorder':'total ascending','tickfont':dict(size=13)}, xaxis={'tickfont':dict(size=13)}, height=700, bargap=0.4, margin=dict(t=20))
        fig2.update_traces(texttemplate='%{x}', textposition='outside', cliponaxis=False)
        st.plotly_chart(fig2, use_container_width=True)

    with col2:
        st.markdown("### Distribuição por tipo de vínculo")
        fig3 = px.pie(vinculo_df, names='Vínculo', values='Servidores', hole=0.45, title=None, color_discrete_sequence=px.colors.qualitative.Pastel)
        fig3.update_layout(legend=dict(orientation="h", y=-0.2, x=0.5, xanchor="center"), height=500, margin=dict(t=20))
        st.plotly_chart(fig3, use_container_width=True)
    # if not filtro_cargo and not filtro_vinculo:
    #     top_vinculo_cargo = df_filtrado.groupby(['vinculo','descricaoCargo']).size().reset_index(name='Servidores')
    #     top_vinculo_cargo = top_vinculo_cargo.sort_values(['Servidores'], ascending=False).groupby('vinculo').head(3)
    #     st.markdown("### Cargos por tipo de vínculo")
    #     fig6 = px.bar(top_vinculo_cargo, x='Servidores', y='vinculo', color='descricaoCargo', orientation='h', title=None, height=600, text='Servidores', color_discrete_sequence=px.colors.sequential.Tealgrn)
    #     fig6.update_traces(texttemplate='%{x}', textposition='outside', cliponaxis=False)
    #     fig6.update_layout(yaxis={'categoryorder':'total ascending','tickfont':dict(size=13)}, xaxis={'tickfont':dict(size=13)}, legend_title_text='Cargo', bargap=0.3, margin=dict(t=20,l=100,r=20))
    #     st.plotly_chart(fig6, use_container_width=True)

## Painel Dados Brutos
elif painel == "Dados Brutos":
    st.title("Dados")
    lotacoes = [s for s in ordem_abas_arquivo if s in dados['descricaoLotacaoPai_subdiretoria'].unique()]
    selecao_lotacoes = st.selectbox("Selecione a lotação para ver os dados:", lotacoes)
    tabela = dados[dados['descricaoLotacaoPai_subdiretoria'] == selecao_lotacoes]
    tabela_formatada = tabela.reset_index(drop=True)
    tabela_formatada = tabela_formatada.drop(columns=['descricaoLotacaoPai_subdiretoria','lotacaoSimplificada'], errors='ignore')
    st.data_editor(tabela_formatada, use_container_width=True, hide_index=True, disabled=True)
    hoje = datetime.today().strftime('%d-%m-%Y')
    nome_arquivo = f'Dados da lotação - {selecao_lotacoes} {hoje}.xlsx'
    if st.button(f"📥 Gerar arquivo da {selecao_lotacoes}"):
        tabela.to_excel(nome_arquivo, index=False, engine='openpyxl')
        with open(nome_arquivo, "rb") as file:
            st.download_button(label=f"⬇️ Baixar dados da {selecao_lotacoes}", data=file,
                file_name=nome_arquivo, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

